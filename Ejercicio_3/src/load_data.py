"""
Load Exercise 3 source data from Excel into SQLite.

The source workbook is preserved without modifications.
Both sheets are loaded into RAW tables so that data-quality
issues can be detected before preparing the information used
for the streak calculation.
"""

from pathlib import Path
import sqlite3

from openpyxl import load_workbook


EXERCISE_DIR = Path(__file__).resolve().parents[1]

EXCEL_PATH = EXERCISE_DIR / "data" / "Rachas.xlsx"
DATABASE_PATH = EXERCISE_DIR / "rachas.db"
SCHEMA_PATH = EXERCISE_DIR / "sql" / "01_schema.sql"

HISTORY_SHEET = "historia"
RETIREMENTS_SHEET = "retiros"


def create_database(connection: sqlite3.Connection) -> None:
    """Create the database schema from the SQL script."""

    schema = SCHEMA_PATH.read_text(encoding="utf-8")
    connection.executescript(schema)


def normalize_date(value) -> str:
    """
    Convert an Excel date/datetime value to ISO format YYYY-MM-DD.

    Raises:
        ValueError: if the value cannot be interpreted as a date.
    """

    if value is None:
        raise ValueError("Date value cannot be null.")

    if hasattr(value, "date"):
        value = value.date()

    try:
        return value.isoformat()
    except AttributeError as exc:
        raise ValueError(f"Invalid date value: {value!r}") from exc


def load_history(
    connection: sqlite3.Connection,
    worksheet,
) -> int:
    """Load the historia worksheet into historia_raw."""

    records = []

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        identificacion, corte_mes, saldo = row

        if identificacion is None:
            raise ValueError(
                f"Missing identificacion in historia row {row_number}."
            )

        if corte_mes is None:
            raise ValueError(
                f"Missing corte_mes in historia row {row_number}."
            )

        if saldo is None:
            raise ValueError(
                f"Missing saldo in historia row {row_number}."
            )

        records.append(
            (
                str(identificacion).strip(),
                normalize_date(corte_mes),
                int(saldo),
            )
        )

    connection.executemany(
        """
        INSERT INTO historia_raw (
            identificacion,
            corte_mes,
            saldo
        )
        VALUES (?, ?, ?)
        """,
        records,
    )

    return len(records)


def load_retirements(
    connection: sqlite3.Connection,
    worksheet,
) -> int:
    """Load the retiros worksheet into retiros_raw."""

    records = []

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        identificacion, fecha_retiro = row

        if identificacion is None:
            raise ValueError(
                f"Missing identificacion in retiros row {row_number}."
            )

        if fecha_retiro is None:
            raise ValueError(
                f"Missing fecha_retiro in retiros row {row_number}."
            )

        records.append(
            (
                str(identificacion).strip(),
                normalize_date(fecha_retiro),
            )
        )

    connection.executemany(
        """
        INSERT INTO retiros_raw (
            identificacion,
            fecha_retiro
        )
        VALUES (?, ?)
        """,
        records,
    )

    return len(records)


def validate_workbook_structure(workbook) -> None:
    """Validate required worksheets and column names."""

    required_sheets = {
        HISTORY_SHEET,
        RETIREMENTS_SHEET,
    }

    missing_sheets = required_sheets - set(workbook.sheetnames)

    if missing_sheets:
        raise ValueError(
            f"Missing required worksheets: {sorted(missing_sheets)}"
        )

    history_headers = [
        cell.value
        for cell in workbook[HISTORY_SHEET][1]
    ]

    retirement_headers = [
        cell.value
        for cell in workbook[RETIREMENTS_SHEET][1]
    ]

    expected_history_headers = [
        "identificacion",
        "corte_mes",
        "saldo",
    ]

    expected_retirement_headers = [
        "identificacion",
        "fecha_retiro",
    ]

    if history_headers != expected_history_headers:
        raise ValueError(
            "Unexpected historia headers. "
            f"Expected {expected_history_headers}, "
            f"received {history_headers}."
        )

    if retirement_headers != expected_retirement_headers:
        raise ValueError(
            "Unexpected retiros headers. "
            f"Expected {expected_retirement_headers}, "
            f"received {retirement_headers}."
        )


def main() -> None:
    """Build the SQLite database from the source Excel workbook."""

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Source workbook not found: {EXCEL_PATH}"
        )

    workbook = load_workbook(
        EXCEL_PATH,
        read_only=True,
        data_only=True,
    )

    validate_workbook_structure(workbook)

    # Rebuild the database to guarantee a reproducible execution.
    if DATABASE_PATH.exists():
        DATABASE_PATH.unlink()

    with sqlite3.connect(DATABASE_PATH) as connection:
        create_database(connection)

        history_count = load_history(
            connection,
            workbook[HISTORY_SHEET],
        )

        retirement_count = load_retirements(
            connection,
            workbook[RETIREMENTS_SHEET],
        )

        connection.commit()

    workbook.close()

    print("Database created successfully.")
    print(f"Database: {DATABASE_PATH}")
    print(f"History records loaded: {history_count}")
    print(f"Retirement records loaded: {retirement_count}")


if __name__ == "__main__":
    main()